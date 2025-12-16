from typing import Any, Dict, Iterable, List, Sequence, Optional

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Servicio de reportes: contratos mínimos para PDF/Excel
# Implementación concreta se hará en iteraciones posteriores.


class ReportService:
    def render_pdf(self, template_name: str, data: Dict[str, Any]) -> bytes:
        """
        Renderiza un PDF a partir de una plantilla HTML/CSS y un diccionario de datos.
        Debe devolver binario del PDF.
        """
        # Placeholder: devolver un PDF válido mínimo para evitar errores de apertura.
        # Este contenido es un PDF muy simple con una página en blanco.
        minimal_pdf = (
            b"%PDF-1.4\n"
            b"1 0 obj<<>>endobj\n"
            b"2 0 obj<< /Type /Catalog /Pages 3 0 R >>endobj\n"
            b"3 0 obj<< /Type /Pages /Kids [4 0 R] /Count 1 >>endobj\n"
            b"4 0 obj<< /Type /Page /Parent 3 0 R /MediaBox [0 0 595 842] >>endobj\n"
            b"xref\n0 5\n0000000000 65535 f \n"
            b"0000000010 00000 n \n"
            b"0000000053 00000 n \n"
            b"0000000106 00000 n \n"
            b"0000000164 00000 n \n"
            b"trailer<< /Root 2 0 R /Size 5 >>\nstartxref\n220\n%%EOF\n"
        )
        return minimal_pdf

    def render_pdf_table(
        self,
        *,
        title: str,
        filters: Dict[str, Any] | None,
        columns: Sequence[str],
        rows: Sequence[Sequence[str | int | None]],
        page_size=A4,
        margin_mm: int = 15,
        accent_hex: str = "#2D3748",
        logo_path: Optional[str] = None,
        landscape_mode: bool = False,
        col_widths_fraction: Optional[Sequence[float]] = None,
    ) -> bytes:
        """Render a simple, robust PDF using ReportLab with a header, filters and a table.
        - Uses system fonts (Helvetica) to avoid external assets.
        - Adds page-friendly margins and alternating row colors.
        """
        buffer = BytesIO()

        # Determine page size (landscape for wide tables)
        ps = landscape(page_size) if landscape_mode else page_size

        doc = SimpleDocTemplate(
            buffer,
            pagesize=ps,
            leftMargin=margin_mm * mm,
            rightMargin=margin_mm * mm,
            topMargin=margin_mm * mm,
            bottomMargin=margin_mm * mm,
        )
        # Ensure pagesize is applied (some viewers cache metadata aggressively)
        try:
            doc.pagesize = ps  # type: ignore[attr-defined]
        except Exception:
            pass

        elements: List[Any] = []

        h_style = ParagraphStyle(name="Heading", fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor(accent_hex))
        n_style = ParagraphStyle(name="Normal", fontName="Helvetica", fontSize=9, leading=11)
        # Ensure wrapping happens within cell width
        try:
            n_style.wordWrap = "LTR"  # type: ignore[attr-defined]
        except Exception:
            pass
        m_style = ParagraphStyle(name="Meta", fontName="Helvetica", fontSize=8, textColor=colors.grey)
        # Header with optional logo (two-column layout to avoid "píxeles piteros")
        # We keep aspect ratio and use a larger, cleaner size.
        if logo_path:
            try:
                logo_img = Image(logo_path)
                # Cap logo size and keep aspect ratio (no distortion)
                # Practical default: ~48x20 mm for decent presence
                max_w, max_h = 48 * mm, 20 * mm
                logo_img._restrictSize(max_w, max_h)  # reportlab safe helper
                logo_img.hAlign = "LEFT"

                header_tbl = Table(
                    [[logo_img, Paragraph(title, h_style)]],
                    colWidths=[max_w, None],
                )
                header_tbl.setStyle(
                    TableStyle(
                        [
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 0),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                            ("TOPPADDING", (0, 0), (-1, -1), 0),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ]
                    )
                )
                elements.append(header_tbl)
            except Exception:
                elements.append(Paragraph(title, h_style))
        else:
            elements.append(Paragraph(title, h_style))
        elements.append(Spacer(1, 4 * mm))

        # Fecha de generación
        from datetime import datetime
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", m_style))

        # Filtros aplicados
        if filters:
            pretty = ", ".join(
                f"{k}: {v}" for k, v in filters.items() if v not in (None, "", [])
            ) or "Sin filtros"
            elements.append(Paragraph(f"Filtros: {pretty}", m_style))

        elements.append(Spacer(1, 5 * mm))

        # Datos de tabla con wrapping: usar Paragraph en celdas para permitir saltos de línea
        data: List[List[Any]] = [list(columns)]
        def _cell(val: Any) -> Any:
            s = "" if val is None else str(val)
            try:
                return Paragraph(s, n_style)
            except Exception:
                return s
        for r in rows:
            data.append([_cell(v) for v in r])

        # Column widths: use caller-provided fractions if present, else even distribution
        if col_widths_fraction:
            total = sum(col_widths_fraction)
            if total <= 0:
                widths = None
            else:
                widths = [doc.width * (w / total) for w in col_widths_fraction]
            table = Table(data, repeatRows=1, colWidths=widths)
        else:
            table = Table(data, repeatRows=1)
            try:
                col_count = len(columns)
                table._argW = [doc.width / col_count] * col_count
            except Exception:
                pass
        table_style = TableStyle([
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(accent_hex)),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("WORDWRAP", (0, 1), (-1, -1), True),
            ("VALIGN", (0, 1), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#EDF2F7")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
        table.setStyle(table_style)

        elements.append(table)

        # Build with simple header/footer functions
        def _header_footer(c: canvas.Canvas, doc_obj: SimpleDocTemplate):
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.grey)
            w, h = ps
            page_num = c.getPageNumber()
            c.drawRightString(w - margin_mm * mm, margin_mm * mm - 6, f"Página {page_num}")
            try:
                c.setTitle(title)
            except Exception:
                pass

        doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)
        return buffer.getvalue()

    def render_excel(self, workbook_spec: Dict[str, Any], rows: Iterable[Dict[str, Any]]) -> bytes:
        """Deprecated placeholder. Use render_excel_xlsx."""
        return b""

    def render_excel_xlsx(
        self,
        *,
        sheet: str,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        date_columns: Sequence[int] | None = None,
        datetime_columns: Sequence[int] | None = None,
        autofilter: bool = True,
        freeze_top_row: bool = True,
    ) -> bytes:
        """Generate a styled XLSX file with header formatting, column widths,
        optional date/datetime formatting, auto-filter and frozen header row.
        """
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = sheet[:31] if sheet else "Hoja1"

        # Header row
        ws.append(list(headers))
        header_fill = PatternFill("solid", fgColor="2D3748")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        # Body rows
        for r in rows:
            ws.append(list(r))

        # Formatting for date/datetime columns (1-based in Excel, but we receive 0-based)
        if date_columns:
            for c0 in date_columns:
                col = c0 + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).number_format = "YYYY-MM-DD"
        if datetime_columns:
            for c0 in datetime_columns:
                col = c0 + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).number_format = "YYYY-MM-DD HH:MM"

        # Auto widths based on content
        for i, _ in enumerate(headers, start=1):
            max_len = 0
            for cell in ws[get_column_letter(i)]:
                val = cell.value
                if val is None:
                    l = 0
                else:
                    l = len(str(val))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

        # Autofilter and freeze pane
        if autofilter:
            ws.auto_filter.ref = ws.dimensions
        if freeze_top_row:
            ws.freeze_panes = "A2"

        # Save to bytes
        from io import BytesIO

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


report_service = ReportService()
